"""Portfolio routes - holdings, transactions, import, MF health."""

import csv
import io

from beanie import PydanticObjectId
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from ....core.constants import SECTOR_MAP
from ....core.response_handler import StandardResponse
from ....core.security import get_current_user
from ....models.documents import Holding
from ....models.documents.holding import EmbeddedTransaction
from ....services.cache import cache_delete
from ....services.portfolio import get_prices_for_holdings, get_user_holdings
from ....services.portfolio.portfolio_service import PortfolioService
from .schemas import (
    HoldingCreate,
    HoldingUpdate,
    ImportResult,
    TransactionCreate,
)

router = APIRouter()


async def _invalidate_portfolio_cache(user_id: str) -> None:
    """Clear all portfolio-related cache keys for a user."""
    for prefix in ("holdings", "sectors", "dashboard", "analytics"):
        await cache_delete(f"{prefix}:{user_id}")


@router.get("", summary="Get portfolio summary")
async def get_portfolio_summary(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    return StandardResponse.ok(await svc.get_summary())


@router.get("/holdings", summary="Get all holdings")
async def get_holdings(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    return StandardResponse.ok(await svc.get_holdings_with_prices())


@router.post("/holdings", summary="Add holding")
async def add_holding(holding: HoldingCreate, current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    try:
        doc_id = await svc.add_holding(
            symbol=holding.symbol,
            name=holding.name,
            exchange=holding.exchange,
            holding_type=holding.holding_type,
            quantity=holding.quantity,
            avg_price=holding.avg_price,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return StandardResponse.ok({"id": doc_id, "symbol": holding.symbol}, "Holding added")


@router.put("/holdings/{holding_id}", summary="Update holding")
async def update_holding(
    holding_id: str, update: HoldingUpdate, current_user: dict = Depends(get_current_user)
) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    try:
        await svc.update_holding(holding_id, quantity=update.quantity, avg_price=update.avg_price)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except LookupError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StandardResponse.ok(message="Holding updated")


@router.delete("/holdings/{holding_id}", summary="Delete holding")
async def delete_holding(holding_id: str, current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    try:
        await svc.delete_holding(holding_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except LookupError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StandardResponse.ok(message="Holding deleted")


@router.get("/sectors", summary="Get sector allocation")
async def get_sector_allocation(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    return StandardResponse.ok(await svc.get_sectors())


@router.get(
    "/dashboard",
    summary="Get dashboard data",
    description="Get complete portfolio dashboard with holdings, sectors, and transactions",
)
async def get_dashboard(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    """Get complete dashboard with holdings, sectors, and recent transactions."""
    from ....services.cache import cache_get, cache_set

    cache_key = f"dashboard:{current_user['_id']}"
    cached = await cache_get(cache_key)
    if cached:
        return StandardResponse.ok(cached)

    holdings = await get_user_holdings(current_user["_id"])
    if not holdings:
        return StandardResponse.ok(
            {
                "holdings": [],
                "sectors": [],
                "xirr": None,
                "transactions": [],
                "summary": {"invested": 0, "current": 0, "pnl": 0, "pnl_pct": 0},
            }
        )

    prices = await get_prices_for_holdings(holdings) or {}
    holdings_list: list = []
    txns: list = []
    total_inv = total_val = 0.0
    sector_values: dict[str, float] = {}

    for h in holdings:
        p = prices.get(h.symbol, {})
        curr_price = p.get("current_price") or h.current_price or h.avg_price
        inv = h.quantity * h.avg_price
        val = h.quantity * curr_price
        total_inv += inv
        total_val += val
        sector = SECTOR_MAP.get(h.symbol, "Others") if h.holding_type != "MF" else h.name
        sector_values[sector] = sector_values.get(sector, 0) + val

        holdings_list.append(
            {
                "_id": str(h.id),
                "symbol": h.symbol,
                "name": h.name,
                "holding_type": h.holding_type,
                "quantity": h.quantity,
                "avg_price": h.avg_price,
                "current_price": round(curr_price, 2),
                "day_change_pct": p.get("day_change_pct", 0),
                "current_value": round(val, 2),
                "total_investment": round(inv, 2),
                "pnl": round(val - inv, 2),
                "pnl_pct": round(((val - inv) / inv * 100) if inv > 0 else 0, 2),
                "sector": sector,
            }
        )
        for i, t in enumerate(h.transactions):
            txns.append({"symbol": h.symbol, "holding_id": str(h.id), "index": i, **t.model_dump()})

    txns.sort(key=lambda x: x.get("date", ""), reverse=True)
    sectors = [
        {"sector": s, "value": round(v, 2), "percentage": round(v / total_val * 100, 1) if total_val > 0 else 0}
        for s, v in sorted(sector_values.items(), key=lambda x: x[1], reverse=True)
    ]

    result = {
        "holdings": holdings_list,
        "sectors": sectors,
        "xirr": _calc_xirr(holdings, total_val),
        "xirr_stocks": _calc_xirr(
            [h for h in holdings if h.holding_type != "MF"],
            sum(hl["current_value"] for hl in holdings_list if hl["holding_type"] != "MF"),
        ),
        "xirr_mf": _calc_xirr(
            [h for h in holdings if h.holding_type == "MF"],
            sum(hl["current_value"] for hl in holdings_list if hl["holding_type"] == "MF"),
        ),
        "transactions": txns[:50],
        "summary": {
            "invested": round(total_inv, 2),
            "current": round(total_val, 2),
            "pnl": round(total_val - total_inv, 2),
            "pnl_pct": round(((total_val - total_inv) / total_inv * 100) if total_inv > 0 else 0, 2),
        },
    }
    await cache_set(cache_key, result, ttl=120 if _market_open() else 3600)
    return StandardResponse.ok(result)


def _market_open() -> bool:
    from ....services.cache import market_open

    return market_open()


def _calc_xirr(holdings, current_value):
    """Calculate XIRR from transaction history using Newton's method."""
    from datetime import datetime

    cashflows = []  # (date, amount) — negative for outflows, positive for inflows
    for h in holdings:
        for t in h.transactions:
            try:
                dt = datetime.strptime(t.date, "%Y-%m-%d")
                amt = t.quantity * t.price
                cashflows.append((dt, -amt if t.type == "BUY" else amt))
            except (ValueError, TypeError):
                continue

    if not cashflows:
        return None

    # Add current portfolio value as final inflow
    cashflows.append((datetime.now(), current_value))
    cashflows.sort(key=lambda x: x[0])

    if len(cashflows) < 2:
        return None

    # Newton's method for XIRR
    d0 = cashflows[0][0]
    days = [(cf[0] - d0).days / 365.0 for cf in cashflows]
    amounts = [cf[1] for cf in cashflows]

    def npv(rate):
        return sum(a / (1 + rate) ** d for a, d in zip(amounts, days))

    def dnpv(rate):
        return sum(-d * a / (1 + rate) ** (d + 1) for a, d in zip(amounts, days))

    rate = 0.1  # initial guess 10%
    for _ in range(100):
        n = npv(rate)
        dn = dnpv(rate)
        if abs(dn) < 1e-12:
            break
        new_rate = rate - n / dn
        if abs(new_rate - rate) < 1e-7:
            rate = new_rate
            break
        rate = new_rate
        if rate < -0.99 or rate > 10:
            return None

    return round(rate * 100, 2) if -1 < rate < 10 else None


@router.get("/transactions", summary="Get transactions")
async def get_transactions(
    page: int = 1, limit: int = 50, current_user: dict = Depends(get_current_user)
) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    return StandardResponse.ok(await svc.get_transactions(page, limit))


@router.post("/transactions", summary="Add transaction")
async def add_transaction(txn: TransactionCreate, current_user: dict = Depends(get_current_user)) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    try:
        result = await svc.add_transaction(
            symbol=txn.symbol,
            txn_type=txn.type,
            quantity=txn.quantity,
            price=txn.price,
            date_str=txn.date.isoformat(),
            holding_type=txn.holding_type,
            notes=txn.notes or "",
            amount=txn.amount,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    msg = "Holding sold completely" if result.get("sold_completely") else "Transaction added"
    return StandardResponse.ok(result, msg)


@router.delete("/transactions/{holding_id}/{index}")
async def delete_transaction(
    holding_id: str, index: int, current_user: dict = Depends(get_current_user)
) -> StandardResponse:
    svc = PortfolioService(current_user["_id"])
    try:
        await svc.delete_transaction(holding_id, index)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except LookupError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StandardResponse.ok(message="Transaction deleted")


@router.post("/import", summary="Import holdings", description="Import holdings from CSV file (Zerodha, Groww)")
async def import_holdings(
    file: UploadFile = File(...), current_user: dict = Depends(get_current_user)
) -> StandardResponse:
    """Import holdings from broker CSV file."""
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files supported")

    text = (await file.read()).decode("utf-8")
    try:
        rows = list(csv.DictReader(io.StringIO(text)))
    except csv.Error as e:
        raise HTTPException(status_code=400, detail=f"Invalid CSV: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Empty CSV")

    headers = [h.lower() for h in rows[0].keys()]
    broker = "zerodha" if "trade_type" in headers else "groww" if "avg cost" in headers else "unknown"
    if broker == "unknown":
        raise HTTPException(status_code=400, detail="Unsupported format")

    user_id = PydanticObjectId(current_user["_id"])
    imported = skipped = 0

    for row in rows:
        try:
            if broker == "zerodha":
                symbol = row.get("symbol", "").split("-")[0].strip().upper()
                qty = float(row.get("quantity", 0))
                price = float(row.get("price", 0))
            else:
                symbol = str(list(row.values())[0]).split()[0].upper()
                qty = float(str(row.get("qty", row.get("quantity", 0))).replace(",", ""))
                price = float(str(row.get("avg cost", row.get("avg_price", 0))).replace(",", ""))

            if not symbol or qty <= 0:
                skipped += 1
                continue

            existing = await Holding.find_one(Holding.user_id == user_id, Holding.symbol == symbol)
            if existing:
                skipped += 1
                continue

            await Holding(user_id=user_id, symbol=symbol, name=symbol, quantity=qty, avg_price=price).insert()
            imported += 1
        except (ValueError, KeyError):
            skipped += 1

    return StandardResponse.ok(ImportResult(broker=broker, imported=imported, skipped=skipped))


@router.post(
    "/import-transactions",
    summary="Import transaction history",
    description="Import transaction history from Groww XLSX (stocks or mutual funds)",
)
async def import_transactions(
    file: UploadFile = File(...), current_user: dict = Depends(get_current_user)
) -> StandardResponse:
    """Import transactions from Groww order history XLSX."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only XLSX files supported")

    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Find header row
    header_row = None
    is_mf = False
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row:
            continue
        vals = [str(v).strip() if v else "" for v in row]
        if "Symbol" in vals:
            header_row = row_idx
            break
        if "Scheme Name" in vals:
            header_row = row_idx
            is_mf = True
            break
    if not header_row:
        raise HTTPException(status_code=400, detail="Unsupported XLSX format")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[header_row]]
    col = {h: i for i, h in enumerate(headers)}
    user_id = PydanticObjectId(current_user["_id"])
    imported = skipped = created = 0

    if is_mf:
        imported, skipped, created = await _import_mf_transactions(ws, header_row, col, user_id)
    else:
        imported, skipped, created = await _import_stock_transactions(ws, header_row, col, user_id)

    return StandardResponse.ok(
        {"imported": imported, "skipped": skipped, "holdings_created": created, "type": "MF" if is_mf else "STOCKS"}
    )


# MF scheme name -> DB symbol mapping
MF_SYMBOL_MAP = {
    "parag parikh flexi cap": "PPFAS",
    "hdfc mid cap": "HDFC-MC",
    "kotak large & midcap": "KOTAK-LM",
    "bandhan small cap": "BANDHAN-SC",
    "pgim india ultra short": "PGIM-USD",
    "axis liquid": "AXIS-LIQ",
    "motilal oswal midcap": "MOTILAL-MC",
    "motilal oswal nifty 200": "MOTILAL-MOM30",
    "jm flexicap": "JM-FLEXI",
    "axis small cap": "AXIS-SC",
    "nippon india small cap": "NIPPON-SC",
}


def _resolve_mf_symbol(scheme_name: str) -> str:
    name = scheme_name.lower()
    for key, symbol in MF_SYMBOL_MAP.items():
        if key in name:
            return symbol
    # Fallback: first 2 words uppercased
    parts = scheme_name.split()[:2]
    return "-".join(p.upper()[:6] for p in parts)


async def _import_mf_transactions(ws, header_row, col, user_id):
    from datetime import datetime

    txn_map = {}  # symbol -> list of txns
    name_map = {}  # symbol -> scheme name

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[col.get("scheme name", 0)]:
            continue
        scheme = str(row[col["scheme name"]]).strip()
        txn_type_raw = str(row[col["transaction type"]]).strip().upper()
        txn_type = "BUY" if txn_type_raw == "PURCHASE" else "SELL" if txn_type_raw == "REDEEM" else None
        if not txn_type:
            continue
        try:
            units = float(row[col["units"]])
            nav = float(row[col["nav"]])
            amount_raw = str(row[col["amount"]]).replace(",", "")
            float(amount_raw)  # validate
            date_str = str(row[col["date"]]).strip()
            dt = datetime.strptime(date_str, "%d %b %Y")
            date_iso = dt.strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            continue

        symbol = _resolve_mf_symbol(scheme)
        if symbol not in txn_map:
            txn_map[symbol] = []
            name_map[symbol] = scheme
        txn_map[symbol].append(
            {"type": txn_type, "quantity": round(units, 4), "price": round(nav, 2), "date": date_iso}
        )

    imported = skipped = created = 0
    for symbol, txns in txn_map.items():
        holding = await Holding.find_one(Holding.user_id == user_id, Holding.symbol == symbol)
        if not holding:
            net_qty = sum(t["quantity"] if t["type"] == "BUY" else -t["quantity"] for t in txns)
            if net_qty <= 0:
                skipped += len(txns)
                continue
            buy_txns = [t for t in txns if t["type"] == "BUY"]
            avg = sum(t["price"] * t["quantity"] for t in buy_txns) / sum(t["quantity"] for t in buy_txns)
            holding = Holding(
                user_id=user_id,
                symbol=symbol,
                name=name_map[symbol],
                quantity=round(net_qty, 4),
                avg_price=round(avg, 2),
                holding_type="MF",
            )
            await holding.insert()
            created += 1

        # Replace all transactions with XLSX (source of truth)
        new_txns = [
            EmbeddedTransaction(type=t["type"], quantity=t["quantity"], price=t["price"], date=t["date"]) for t in txns
        ]
        holding.transactions = new_txns
        buys = [t for t in new_txns if t.type == "BUY"]
        sells = [t for t in new_txns if t.type == "SELL"]
        buy_qty = sum(t.quantity for t in buys)
        sell_qty = sum(t.quantity for t in sells)
        holding.quantity = round(buy_qty - sell_qty, 4)
        if buys:
            holding.avg_price = round(sum(t.quantity * t.price for t in buys) / buy_qty, 2)
        holding.holding_type = "MF"
        await holding.save()
        imported += len(txns)

    return imported, skipped, created


async def _import_stock_transactions(ws, header_row, col, user_id):
    from collections import Counter, defaultdict
    from datetime import datetime

    required = {"symbol", "type", "quantity", "value", "execution date and time"}
    if not required.issubset(col.keys()):
        raise HTTPException(status_code=400, detail=f"Missing columns: {required - col.keys()}")

    txn_map = defaultdict(list)
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[col["symbol"]]:
            continue
        symbol = str(row[col["symbol"]]).strip().upper()
        txn_type = str(row[col["type"]]).strip().upper()
        if txn_type not in ("BUY", "SELL"):
            skipped += 1
            continue
        try:
            qty = float(row[col["quantity"]])
            value = float(row[col["value"]])
            price = round(value / qty, 2) if qty > 0 else 0
            date_str = str(row[col["execution date and time"]]).strip()
            dt = datetime.strptime(date_str, "%d-%m-%Y %I:%M %p")
            date_iso = dt.strftime("%Y-%m-%d")
        except (ValueError, ZeroDivisionError):
            skipped += 1
            continue

        name = str(row[col.get("stock name", col.get("symbol"))]).strip()
        txn_map[symbol].append({"type": txn_type, "quantity": qty, "price": price, "date": date_iso, "name": name})

    imported = created = 0
    for symbol, txns in txn_map.items():
        holding = await Holding.find_one(Holding.user_id == user_id, Holding.symbol == symbol)
        if not holding:
            net_qty = sum(t["quantity"] if t["type"] == "BUY" else -t["quantity"] for t in txns)
            if net_qty <= 0:
                skipped += len(txns)
                continue
            buy_txns = [t for t in txns if t["type"] == "BUY"]
            avg = sum(t["price"] * t["quantity"] for t in buy_txns) / sum(t["quantity"] for t in buy_txns)
            holding = Holding(
                user_id=user_id,
                symbol=symbol,
                name=txns[0]["name"],
                quantity=net_qty,
                avg_price=round(avg, 2),
            )
            await holding.insert()
            created += 1

        # Use Counter-based dedup to handle same-day trades
        xlsx_counter = Counter((t["date"], t["type"], t["quantity"]) for t in txns)
        db_counter = Counter((t.date, t.type, t.quantity) for t in holding.transactions)
        missing = xlsx_counter - db_counter

        added = 0
        for (date, ttype, qty), count in missing.items():
            matching = [t for t in txns if t["date"] == date and t["type"] == ttype and t["quantity"] == qty]
            for i in range(count):
                if i < len(matching):
                    t = matching[i]
                    holding.transactions.append(
                        EmbeddedTransaction(type=t["type"], quantity=t["quantity"], price=t["price"], date=t["date"])
                    )
                    added += 1

        if added:
            buys = [t for t in holding.transactions if t.type == "BUY"]
            sells = [t for t in holding.transactions if t.type == "SELL"]
            buy_qty = sum(t.quantity for t in buys)
            sell_qty = sum(t.quantity for t in sells)
            holding.quantity = buy_qty - sell_qty
            if buys:
                holding.avg_price = round(sum(t.quantity * t.price for t in buys) / buy_qty, 2)
            await holding.save()
            imported += added
        else:
            skipped += len(txns)

    return imported, skipped, created


def categorize_mf(name: str) -> tuple[str, int]:
    """Categorize mutual fund by name and return expected benchmark return."""
    name = name.upper()
    if any(x in name for x in ["LIQUID", "OVERNIGHT"]):
        return "Liquid", 6
    if any(x in name for x in ["DEBT", "BOND"]):
        return "Debt", 7
    if "SMALL" in name:
        return "Small Cap", 15
    if "MID" in name:
        return "Mid Cap", 14
    if any(x in name for x in ["LARGE", "BLUECHIP"]):
        return "Large Cap", 12
    if any(x in name for x in ["INDEX", "NIFTY"]):
        return "Index", 12
    return "Equity", 12


@router.get("/mf/health", summary="MF health check", description="Analyze mutual fund performance and health")
async def mf_health_check(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    """Analyze mutual fund portfolio health with benchmarks and recommendations."""
    from ....services.cache import cache_get, cache_set

    ck = f"mf_health:{current_user['_id']}"
    cached = await cache_get(ck)
    if cached:
        return StandardResponse.ok(cached)

    holdings = await get_user_holdings(current_user["_id"])
    mf_holdings = [h for h in holdings if h.holding_type == "MF"]

    if not mf_holdings:
        return StandardResponse.ok(
            {"message": "No mutual funds in portfolio", "funds": [], "total_mf_value": 0, "health_score": 100}
        )

    prices = await get_prices_for_holdings(mf_holdings) or {}
    analysis = []
    issues = []
    total_expense = 0
    total_value = 0

    for h in mf_holdings:
        name = (h.name or "").upper()
        curr_price = prices.get(h.symbol, {}).get("current_price") or h.current_price or h.avg_price
        value = h.quantity * curr_price
        total_value += value

        # Estimate expense ratio
        if "DIRECT" in name:
            expense_ratio = 0.5
        elif "INDEX" in name or "ETF" in name:
            expense_ratio = 0.2
        else:
            expense_ratio = 1.5

        annual_expense = value * expense_ratio / 100
        total_expense += annual_expense

        # Categorize and get benchmark
        if "LIQUID" in name or "OVERNIGHT" in name or "MONEY" in name:
            category, benchmark_return = "Liquid", 6
        elif "DEBT" in name or "BOND" in name or "GILT" in name:
            category, benchmark_return = "Debt", 7
        elif "SMALL" in name:
            category, benchmark_return = "Small Cap", 15
        elif "MID" in name:
            category, benchmark_return = "Mid Cap", 14
        elif "LARGE" in name or "BLUECHIP" in name:
            category, benchmark_return = "Large Cap", 12
        elif "FLEXI" in name or "MULTI" in name:
            category, benchmark_return = "Flexi Cap", 13
        elif "INDEX" in name or "NIFTY" in name or "SENSEX" in name:
            category, benchmark_return = "Index", 12
        elif "INTERNATIONAL" in name or "US" in name or "GLOBAL" in name or "NASDAQ" in name:
            category, benchmark_return = "International", 14
        else:
            category, benchmark_return = "Equity", 12

        # Calculate returns
        invested = h.quantity * h.avg_price
        returns_pct = ((value - invested) / invested * 100) if invested > 0 else 0

        # Status
        if returns_pct < benchmark_return - 5:
            status = "Underperforming"
            issues.append(f"{h.symbol} is underperforming benchmark by {round(benchmark_return - returns_pct, 1)}%")
        elif returns_pct > benchmark_return + 5:
            status = "Outperforming"
        else:
            status = "On Track"

        analysis.append(
            {
                "symbol": h.symbol,
                "name": h.name,
                "category": category,
                "value": round(value, 2),
                "returns_pct": round(returns_pct, 2),
                "benchmark_return": benchmark_return,
                "expense_ratio": expense_ratio,
                "annual_expense": round(annual_expense, 2),
                "status": status,
            }
        )

    # Check overlap
    categories = [a["category"] for a in analysis]
    category_counts = {c: categories.count(c) for c in set(categories)}
    for cat, count in category_counts.items():
        if count > 2:
            issues.append(f"High overlap: {count} funds in {cat} category")

    # Check expense
    avg_expense = (total_expense / total_value * 100) if total_value > 0 else 0
    if avg_expense > 1:
        issues.append(f"High expense ratio: {round(avg_expense, 2)}% - Consider switching to direct plans")

    # Recommendations
    recommendations = []
    underperformers = [a for a in analysis if a["status"] == "Underperforming"]
    if underperformers:
        recommendations.append(
            {
                "type": "switch",
                "message": f"Consider switching {len(underperformers)} underperforming fund(s)",
                "funds": [u["symbol"] for u in underperformers],
            }
        )

    high_expense = [a for a in analysis if a["expense_ratio"] > 1]
    if high_expense:
        recommendations.append(
            {
                "type": "expense",
                "message": "Switch to direct plans to save on expense ratio",
                "potential_savings": round(sum(a["annual_expense"] * 0.5 for a in high_expense), 2),
            }
        )

    if not any("Index" in a["category"] for a in analysis):
        recommendations.append({"type": "add", "message": "Consider adding low-cost index funds for core allocation"})

    result = {
        "total_mf_value": round(total_value, 2),
        "total_annual_expense": round(total_expense, 2),
        "avg_expense_ratio": round(avg_expense, 2),
        "funds": analysis,
        "issues": issues,
        "health_score": max(10, 100 - len(issues) * 10),
        "recommendations": recommendations,
        "note": "Returns comparison is vs annual benchmarks. Short holding periods may show underperformance.",
    }
    await cache_set(ck, result, ttl=600)
    return StandardResponse.ok(result)


@router.get("/mf/overlap", summary="MF overlap analysis", description="Analyze stock overlap between mutual funds")
async def mf_overlap(current_user: dict = Depends(get_current_user)) -> StandardResponse:
    """Analyze overlap between mutual funds."""
    holdings = await get_user_holdings(current_user["_id"])
    mf_holdings = [h for h in holdings if h.holding_type == "MF"]
    if len(mf_holdings) < 2:
        return StandardResponse.ok({"overlaps": [], "message": "Need at least 2 MFs for overlap analysis"})

    return StandardResponse.ok({"overlaps": [], "total_overlap_pct": 0, "message": "Overlap data not available"})


@router.get("/mf/expense-impact", summary="MF expense impact", description="Calculate long-term expense ratio impact")
async def mf_expense_impact(years: int = 20, current_user: dict = Depends(get_current_user)) -> StandardResponse:
    """Calculate long-term impact of expense ratios with direct plan comparison."""
    holdings = await get_user_holdings(current_user["_id"])
    mf_holdings = [h for h in holdings if h.holding_type == "MF"]
    if not mf_holdings:
        return StandardResponse.ok({"current_value": 0, "potential_savings": 0, "message": "No mutual funds"})

    prices = await get_prices_for_holdings(mf_holdings)
    total_value = 0
    total_expense_current = 0
    total_expense_direct = 0

    for h in mf_holdings:
        name = (h.name or "").upper()
        curr = prices.get(h.symbol, {}).get("current_price") or h.current_price or h.avg_price
        value = h.quantity * curr
        total_value += value

        # Current vs direct expense ratios
        if "DIRECT" in name:
            current_exp, direct_exp = 0.5, 0.5
        elif "INDEX" in name:
            current_exp, direct_exp = 0.2, 0.1
        else:
            current_exp, direct_exp = 1.5, 0.5

        total_expense_current += value * current_exp / 100
        total_expense_direct += value * direct_exp / 100

    # Project with 12% return
    def project_value(principal, expense_ratio, years):
        value = principal
        for _ in range(years):
            value = value * (1 + 0.12 - expense_ratio / 100)
        return value

    current_expense_ratio = (total_expense_current / total_value * 100) if total_value > 0 else 0
    direct_expense_ratio = (total_expense_direct / total_value * 100) if total_value > 0 else 0

    future_current = project_value(total_value, current_expense_ratio, years)
    future_direct = project_value(total_value, direct_expense_ratio, years)

    return StandardResponse.ok(
        {
            "current_value": round(total_value, 2),
            "current_expense_ratio": round(current_expense_ratio, 2),
            "direct_expense_ratio": round(direct_expense_ratio, 2),
            "annual_expense_current": round(total_expense_current, 2),
            "annual_expense_direct": round(total_expense_direct, 2),
            "projection_years": years,
            "future_value_current": round(future_current, 2),
            "future_value_direct": round(future_direct, 2),
            "potential_savings": round(future_direct - future_current, 2),
            "message": (
                f"Switching to direct plans could save you "
                f"₹{round((future_direct - future_current)/100000, 1)}L over {years} years"
            ),
        }
    )


@router.get(
    "/earnings-calendar",
    summary="Earnings calendar",
    description="Upcoming earnings for user's holdings",
)
async def get_earnings_calendar(
    current_user: dict = Depends(get_current_user),
) -> StandardResponse:
    """Get upcoming earnings dates for user's stock holdings."""
    import httpx

    from ....services.cache import cache_get, cache_set

    ck = f"earnings_cal:{current_user['_id']}"
    cached = await cache_get(ck)
    if cached:
        return StandardResponse.ok(cached)

    holdings = await get_user_holdings(current_user["_id"])
    equity = [h for h in holdings if h.holding_type != "MF"]
    if not equity:
        return StandardResponse.ok({"earnings": []})

    earnings = []

    import asyncio

    async def _fetch_earnings(symbol):
        try:
            async with httpx.AsyncClient(timeout=8) as c:
                # Try Yahoo Finance first
                resp = await c.get(
                    f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}.NS",
                    headers={"User-Agent": "Mozilla/5.0"},
                )
                if resp.status_code == 200:
                    meta = resp.json().get("chart", {}).get("result", [{}])[0].get("meta", {})
                    ts = meta.get("earningsTimestamp")
                    if ts:
                        from datetime import datetime, timezone

                        dt = datetime.fromtimestamp(ts, tz=timezone.utc)
                        if dt > datetime.now(timezone.utc):
                            return {
                                "symbol": symbol,
                                "date": dt.isoformat(),
                                "date_str": dt.strftime("%d %b %Y"),
                                "source": "yahoo",
                            }

                # Fallback to Screener.in
                screener_resp = await c.get(
                    f"https://www.screener.in/company/{symbol}/",
                    headers={"User-Agent": "Mozilla/5.0"},
                    follow_redirects=True,
                )
                if screener_resp.status_code == 200:
                    import re

                    text = screener_resp.text
                    # Look for "Results" or "Earnings" date patterns
                    date_pattern = (
                        r"(?:Results|Earnings|Board Meeting).*?"
                        r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})"
                    )
                    match = re.search(date_pattern, text, re.IGNORECASE)
                    if match:
                        from datetime import datetime, timezone

                        try:
                            dt = datetime.strptime(match.group(1), "%d %b %Y").replace(tzinfo=timezone.utc)
                            if dt > datetime.now(timezone.utc):
                                return {
                                    "symbol": symbol,
                                    "date": dt.isoformat(),
                                    "date_str": dt.strftime("%d %b %Y"),
                                    "source": "screener",
                                }
                        except ValueError:
                            pass
        except Exception:
            pass
        return None

    results = await asyncio.gather(*[_fetch_earnings(h.symbol) for h in equity[:20]])
    earnings = [r for r in results if r]

    earnings.sort(key=lambda x: x["date"])
    result = {"earnings": earnings}
    await cache_set(ck, result, ttl=3600)
    return StandardResponse.ok(result)


@router.get("/snapshots", summary="Portfolio growth over time")
async def get_snapshots(range: str = "3M", current_user: dict = Depends(get_current_user)) -> StandardResponse:
    from datetime import datetime, timedelta

    from ....models.documents.portfolio_snapshot import PortfolioSnapshot
    from ....services.cache import cache_get, cache_set

    ck = f"snapshots:{current_user['_id']}:{range}"
    cached = await cache_get(ck)
    if cached:
        return StandardResponse.ok(cached)

    uid = PydanticObjectId(current_user["_id"])
    days = {"1W": 7, "1M": 30, "3M": 90, "6M": 180, "1Y": 365, "ALL": 3650}.get(range, 90)
    since = datetime.utcnow() - timedelta(days=days)

    snaps = (
        await PortfolioSnapshot.find(PortfolioSnapshot.user_id == uid, PortfolioSnapshot.date >= since)
        .sort("+date")
        .to_list()
    )

    data = [
        {
            "date": s.date.strftime("%Y-%m-%d"),
            "value": round(s.value),
            "invested": round(s.invested),
            "pnl": round(s.pnl),
            "pnl_pct": round(s.pnl_pct, 1),
        }
        for s in snaps
    ]
    result = {"snapshots": data, "range": range, "count": len(data)}
    await cache_set(ck, result, ttl=3600)
    return StandardResponse.ok(result)
